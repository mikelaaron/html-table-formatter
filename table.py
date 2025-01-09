from openpyxl import load_workbook

def generate_html_table(excel_file, sheet_name):
    """
    Generates an HTML table from an Excel spreadsheet.

    Args:
        excel_file: Path to the Excel file.
        sheet_name: Name of the sheet containing the data.

    Returns:
        HTML table code as a string.
    """

    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]

    html_table = """
    <table style="border-collapse: collapse; width: 100%;">
      <thead>
        <tr>
          <th style="border: 1px solid black; padding: 8px; text-align: center; width: 20%; color: #EB3F43;"><strong>Fall Highlights</strong></th>
          <th style="border: 1px solid black; padding: 8px; text-align: center; width: 55%; color: #EB3F43;"><strong>Details</strong></th>
          <th style="border: 1px solid black; padding: 8px; text-align: center; width: 25%; color: #EB3F43;"><strong>Date</strong></th>
        </tr>
      </thead>
      <tbody>
    """

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip headers
        title, details, date = row
        html_table += f"""
        <tr>
          <td style="border: 1px solid black; padding: 8px; text-align: left; vertical-align: top;"><strong>{title}</strong></td>
          <td style="border: 1px solid black; padding: 8px; text-align: left;">{details}</td>
          <td style="border: 1px solid black; padding: 8px; text-align: left;">{date}</td>
        </tr>
        """

    html_table += """
      </tbody>
    </table>
    """
    return html_table

# Example usage:
excel_file = "/Users/michael.aaron/Library/CloudStorage/OneDrive-Ogilvy/DESKTOP/WinterPost.xlsx"
sheet_name = "Sheet1"  

html_output = generate_html_table(excel_file, sheet_name)
print(html_output)

# You can then copy this output and paste it into your newsletter editor.